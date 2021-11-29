from bs4 import BeautifulSoup
import requests
import openpyxl
import json
import math

# Excel that que contains information of 40000 movies
excel_document = openpyxl.load_workbook('MovieGenreIGC_v3.xlsx')
page = excel_document.active
url_list = []
plot_url_list = []
cast_url_list = []
keywords_url_list = []
url_http = "http://www.imdb.com/title/tt"
plot_to_go = "/plotsummary?ref_=tt_ov_pl"
full_cast_to_go = "/fullcredits?ref_=tt_ov_st_sm"
keywords_to_go = "/keywords?ref_=tt_stry_kw"


# Variables para crear nuestro archivo JSON
file = open("jsonFilms.json", "a")
# with open("jsonFilms.json") as f:
#     numberOfJson = sum(1 for line in f)
id = 138
# print(numberOfJson)
# # Contador que comienza en la primera línea del archivo excel
# if(numberOfJson == 0):
#     counter = 2
# else:
#     counter = math.ceil(numberOfJson/3)

for i in range(140, (page.max_row + 1)):
    seed = page.cell(row = i, column = 1).value
    full_seed = str(seed).zfill(7)
    full_url = url_http + full_seed
    if(full_url not in url_list): 
        url_list.append(full_url)
    

for i in range(0, len(url_list)):
    headerString = {
        "index": {"_index": "films", "_id": id}
    }

    file.write(json.dumps(headerString))
    file.write("\n")
    screenwriters_list = []
    genres_list = []
    language_list = []
    cast_list = []
    keywords_list = []
    
    plot_url_list.append(url_list[i] + plot_to_go)
    cast_url_list.append(url_list[i] + full_cast_to_go)
    keywords_url_list.append(url_list[i] + keywords_to_go)
    response = requests.get(url_list[i])
    response2 = requests.get(plot_url_list[i])
    response3 = requests.get(cast_url_list[i])
    respones4 = requests.get(keywords_url_list[i])
    soup = BeautifulSoup(response.text, 'html.parser')
    plot_soup = BeautifulSoup(response2.text, 'html.parser')
    cast_soup = BeautifulSoup(response3.text, 'html.parser')
    keywords_soup = BeautifulSoup(respones4.text, 'html.parser')

    # Obtención del título en castellano
    title = soup.find('h1', attrs={'data-testid': 'hero-title-block__title'}).get_text()
    # Obtención del título original (en su idioma original)
    original_title = soup.find('div', attrs={'data-testid': 'hero-title-block__original-title'})

    # Obtención de los generos de cada uno de las películas
    genres = soup.find('div', attrs={'data-testid': 'genres'}).find_all('span', attrs={'class': 'ipc-chip__text'})
  
    # Obtención del año de estreno de cada una de las películas
    year = soup.find('span', attrs={'class': 'TitleBlockMetaData__ListItemText-sc-12ein40-2'}).get_text()

    #Obtención del rating de la película
    rating = soup.find('span', attrs={'class' :'AggregateRatingButton__RatingScore-sc-1ll29m0-1'}).get_text()

    #Obtemción del argumento de la pelicula (lo consultamos en una página a la que se nos redirecciona)
    plot = plot_soup.find('ul', attrs={'id' :'plot-summaries-content'}).find('p').get_text()

    # Obtención del reparto completo de la película (todos sus actores)
    if((cast_soup.find('table', attrs={'class': 'cast_list'})) is not None):
        cast = cast_soup.find('table', attrs={'class': 'cast_list'}).find_all('td', attrs={'class': ''})
    else:
        cast = []
    # Obtención del director/es de la película
    director = cast_soup.find_all('table', attrs={'class' : 'simpleCreditsTable'})[0].find('td', attrs={'class' : 'name'}).get_text().replace("\n", "")

    #Obtencion de los guionistas de la pelicula
    screenwriters = cast_soup.find_all('table', attrs={'class' : 'simpleCreditsTable'})[1].find_all('td', attrs={'class' : 'name'})

    #Palabras clave de la pelicula
    if((keywords_soup.find('table', attrs={'class' : 'evenWidthTable2Col'})) is not None):
        keywords = keywords_soup.find('table', attrs={'class' : 'evenWidthTable2Col'}).find_all('div', attrs={'class' : 'sodatext'})
    else:
        keywords = []
    #idioma de la película
    language = soup.find('li', attrs={'data-testid': 'title-details-languages'}).find_all('a', attrs={'class': 'ipc-metadata-list-item__list-content-item ipc-metadata-list-item__list-content-item--link'})
    
    # Mostramos todos los datos por pantalla para mostrar
    for i in range(len(language)):
        if(language[i].get_text() != "None"):
            language_list.append(language[i].get_text())

    for i in range(len(genres)):
        genres_list.append(genres[i].get_text())

    for i in range(len(cast)):
        cast_list.append(cast[i].get_text().replace("\n", ""))

    for i in range(len(screenwriters)):
        screenwriters_list.append(screenwriters[i].get_text().replace("\n", ""))

    
    for i in range(len(keywords)):
        keywords_list.append(keywords[i].get_text().replace("\n", ""))

    if(original_title is not None):
        original_title = original_title.get_text().split(': ')
        jsonString = {
            "title": title,
            "originalTitle": original_title[1],
            "genres": genres_list,
            "year": year,
            "rating": rating,
            "plot": plot,
            "director": director,
            "cast": cast_list,
            "screenwriters": screenwriters_list,
            "keywords": keywords_list,
            "language": language_list
        }
        
        file.write(json.dumps(jsonString))
        file.write("\n\n")
    else:
        jsonString = {
            "title": title,
            "originalTitle": "Ninguno",
            "genres": genres_list,
            "year": year,
            "rating": rating,
            "plot": plot,
            "director": director,
            "cast": cast_list,
            "screenwriters": screenwriters_list,
            "keywords": keywords_list,
            "language": language_list
         }
    
        file.write(json.dumps(jsonString))
        file.write("\n\n")
    
    print('----PELÍCULA--------',id,'-----------INSERTADA EN EL JSON------------')
    id += 1 
    


file.close()
    
